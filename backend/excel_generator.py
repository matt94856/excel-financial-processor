import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List, Any
import logging
from pathlib import Path

logger = logging.getLogger(__name__)

class ExcelGenerator:
    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.right_alignment = Alignment(horizontal='right', vertical='center')
    
    def generate_excel(self, data: Dict[str, Any], output_path: Path):
        """Generate formatted Excel file from processed data"""
        try:
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Add summary sheet
            self._create_summary_sheet(wb, data)
            
            # Add estimate sheets
            for estimate in data['estimates']:
                self._create_estimate_sheet(wb, estimate)
            
            # Add financial statement sheets
            for financial in data['financial_statements']:
                self._create_financial_sheet(wb, financial)
            
            # Add raw data sheets
            for sheet_name, sheet_data in data['sheets'].items():
                self._create_raw_data_sheet(wb, sheet_name, sheet_data)
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"Excel file generated successfully: {output_path}")
            
        except Exception as e:
            logger.error(f"Error generating Excel file: {str(e)}")
            raise
    
    def _create_summary_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create summary sheet"""
        ws = wb.create_sheet("Summary")
        
        # Add title
        ws['A1'] = "FINANCIAL DOCUMENT PROCESSOR - SUMMARY"
        ws['A1'].font = Font(bold=True, size=16, color="366092")
        ws.merge_cells('A1:D1')
        
        # Add file information
        ws['A3'] = f"Source File: {data['file_name']}"
        ws['A3'].font = Font(bold=True)
        
        # Add summary data
        row = 5
        summary_data = [
            ("Total Estimates", data['summary']['total_estimates']),
            ("Total Financial Statements", data['summary']['total_financial_statements']),
            ("Total Sheets Processed", data['summary']['total_sheets']),
            ("Grand Total", f"${data['summary']['grand_total']:.2f}")
        ]
        
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            if row == len(summary_data) + 4:  # Grand total row
                ws[f'B{row}'].font = Font(bold=True, size=14)
            row += 1
        
        # Format summary sheet
        self._format_summary_sheet(ws)
    
    def _create_estimate_sheet(self, wb: Workbook, estimate: Dict[str, Any]):
        """Create estimate sheet"""
        ws = wb.create_sheet(f"Estimate_{estimate['sheet_name']}")
        
        # Add title
        ws['A1'] = estimate['title']
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        ws.merge_cells('A1:D1')
        
        # Add headers
        headers = ['Description', 'Quantity', 'Unit Price', 'Total']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_alignment
        
        # Add data rows
        row = 4
        for item in estimate['items']:
            ws.cell(row=row, column=1, value=item['description'])
            ws.cell(row=row, column=2, value=item['quantity'])
            ws.cell(row=row, column=3, value=item['unit_price'])
            ws.cell(row=row, column=4, value=item['total'])
            
            # Format row
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                if col in [2, 3, 4]:  # Numeric columns
                    cell.alignment = self.right_alignment
                    cell.number_format = '#,##0.00'
            
            row += 1
        
        # Add total row
        ws.cell(row=row, column=3, value="TOTAL:")
        ws.cell(row=row, column=4, value=estimate['total'])
        
        # Format total row
        for col in range(3, 5):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.border = self.border
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            if col == 4:
                cell.alignment = self.right_alignment
                cell.number_format = '#,##0.00'
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _create_financial_sheet(self, wb: Workbook, financial: Dict[str, Any]):
        """Create financial statement sheet"""
        ws = wb.create_sheet(f"Financial_{financial['sheet_name']}")
        
        # Add title
        ws['A1'] = financial['title']
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        ws.merge_cells('A1:B1')
        
        row = 3
        
        # Process sections
        for section in financial['sections']:
            # Section header
            ws.cell(row=row, column=1, value=section['name'])
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)
            row += 1
            
            # Section items
            if section['items']:
                # Headers
                ws.cell(row=row, column=1, value="Description")
                ws.cell(row=row, column=2, value="Amount")
                
                for col in range(1, 3):
                    cell = ws.cell(row=row, column=col)
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.border = self.border
                    cell.alignment = self.center_alignment
                
                row += 1
                
                # Items
                for item in section['items']:
                    ws.cell(row=row, column=1, value=item['description'])
                    ws.cell(row=row, column=2, value=item['amount'])
                    
                    # Format row
                    for col in range(1, 3):
                        cell = ws.cell(row=row, column=col)
                        cell.border = self.border
                        if col == 2:  # Amount column
                            cell.alignment = self.right_alignment
                            cell.number_format = '#,##0.00'
                    
                    row += 1
            
            row += 1  # Space between sections
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _create_raw_data_sheet(self, wb: Workbook, sheet_name: str, sheet_data: Dict[str, Any]):
        """Create raw data sheet"""
        ws = wb.create_sheet(f"Raw_{sheet_name}")
        
        # Add title
        ws['A1'] = sheet_data['title']
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        
        # Add headers
        if 'headers' in sheet_data:
            for col, header in enumerate(sheet_data['headers'], 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = self.center_alignment
        
        # Add data
        if 'data' in sheet_data:
            row = 4
            for record in sheet_data['data']:
                for col, (key, value) in enumerate(record.items(), 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = self.border
                row += 1
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _format_summary_sheet(self, ws):
        """Format summary sheet"""
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        
        # Add borders to data area
        for row in range(5, 9):
            for col in range(1, 3):
                ws.cell(row=row, column=col).border = self.border
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = None
            
            # Find the first non-merged cell to get column letter
            for cell in column:
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                    break
            
            if column_letter is None:
                continue  # Skip if we can't determine column letter
            
            for cell in column:
                try:
                    if hasattr(cell, 'value') and cell.value is not None:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
